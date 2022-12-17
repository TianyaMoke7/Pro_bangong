# 计划部使用软件

from ast import Break
import collections
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook

# from ast import While
# import openpyxl
from openpyxl.styles import Font,colors, Alignment
# from openpyxl import Workbook
from openpyxl.styles import *
# # from  openpyxl import  Workbook
# import xlrd
# import xlwt
# from pathlib import Path
import os
# import shutil

# from xlwt import Workbook
# import string
# import tkinter as tk
# from tkinter import W, filedialog

# 摘抄的引用
from datetime import datetime, time,timedelta,date
# from .models import (hostIDTable,equipmentResume,
# productionResume,problemFeedback,feddbackCase,summaryproduction,summary_hostIDTable,userid_name)

# Search_Path = 'C:\\Users\\wu_jianguo\\Desktop\\请购计划'   # 计划应先将当天的“生产订单在制物料分析表”、“货位存量查询”放入此路径下
Search_Path = filedialog.askdirectory()   # 计划应先将当天的“生产订单在制物料分析表”、“货位存量查询”放入此路径下

Path_Plan = '\\\\172.16.1.52\JHT-design\\JHT-CKP-CIP-文件管理中心\\生产部\\00-生产计划\\JHT 计划部部门工作\\JHT产线需求清单列表\\产线需求'  # “计划部请购大表”
Path_Purchase = '\\\\172.16.1.52\\JHT-design\\JHT-CKP-CIP-文件管理中心\\采购部\\采购对外文件\\对外订单列表\\JHT账套'  #采购更新的“JHT共享采购订单列表”
Files_Search_Path = os.listdir(Search_Path)
Files_Path_Plan = os.listdir(Path_Plan)
Files_Path_Purchase = os.listdir(Path_Purchase)

# 打开几张大表
for plan_file in Files_Path_Plan:
    if plan_file[0:7] == '计划部请购大表':
        print('正在打开'+' “'+plan_file+'” '+'请稍等...')
        break
excel_Plan_PurchaseBigTable_path = Path_Plan + '\\' + plan_file
# excel_Plan_PurchaseBigTable_path = 'C:\\Users\\wu_jianguo\\Desktop\\请购计划\\计划部请购大表（量产+新品）--2022-7-27（仅设变）.xlsx'
wb = load_workbook(excel_Plan_PurchaseBigTable_path)
sheet_NewPlan = wb.worksheets[0]
if sheet_NewPlan.title[4] == '月':
    month_need = sheet_NewPlan.title[3] + '月份需求数量'
else:
    month_need = sheet_NewPlan.title[3:5] + '月份需求数量'
print('正在打开'+' “'+plan_file+'” '+'请稍等...')
# excel_Plan_PurchaseBigTable = pd.read_excel(excel_Plan_PurchaseBigTable_path,sheet_name=0,header=3)
excel_Plan_PurchaseBigTable = pd.read_excel(excel_Plan_PurchaseBigTable_path,sheet_name=0,header=3,usecols=["子件编码","费用","替代料1","替代料2",month_need])
# excel_Plan_PurchaseBigTable = pd.read_excel(r'C:\Users\wu_jianguo\Desktop\请购计划\新建 XLSX 工作表.xlsx',sheet_name=0,header=3,usecols="D,H:J,CB")

for purchase_file in Files_Path_Purchase:
    if purchase_file[0:11] == 'JHT共享采购订单列表':
        print('正在打开'+' “'+purchase_file+'” '+'请稍等...')
        break
# excel_PurchaseListTable_path = Path_Purchase + '\\' + purchase_file
excel_PurchaseListTable_path = '\\\\172.16.1.52\\JHT-design\\JHT-CKP-CIP-文件管理中心\\采购部\\采购对外文件\\对外订单列表\\JHT账套\\JHT共享采购订单列表.XLSX'
excel_PurchaseListTable = pd.read_excel(excel_PurchaseListTable_path,sheet_name=0,header=0,usecols=["存货编码","未入库量","行关闭人"])
# excel_PurchaseListTable = pd.read_excel(r'C:\Users\wu_jianguo\Desktop\请购计划\JHT共享采购订单列表.xlsx',sheet_name=0,header=0,usecols=["关闭人","存货编码","未入库量","行关闭人"])

for making_file in Files_Search_Path:
    if making_file[0:11] == '生产订单在制物料分析表':
        print('正在打开'+' “'+making_file+'” '+'请稍等...')
        break
excel_MakingTable_path = Search_Path + '/' + making_file
excel_MakingTable = pd.read_excel(excel_MakingTable_path,sheet_name=0,header=5,usecols=["子件物料编码","应领数量","已领数量"])

for stocks_file in Files_Search_Path:
    if stocks_file[0:6] == '货位存量查询':
        print('正在打开'+' “'+stocks_file+'” '+'请稍等...')
        break
excel_StocksTable_path = Search_Path + '/' + stocks_file
excel_StocksTable = pd.read_excel(excel_StocksTable_path,sheet_name=0,header=1,usecols=["存货编码","仓库名称","现存数量"])

Sum_shortage = {}    # 缺料数量  shortage = plan - over
Sum_plan = {}        # 计划需求总数量
Branch = tuple()     # 单行替代料
Branch_List = list() # 所有替代料
Sum_purchase = {}    # 采购在途数量
Sum_making = {}      # 在制数量
Sum_stocks = {}      # 库存数量
Sum_over = {}        # over = purchase + making + stocks

Branch1_List = list()# 请购大表的物料数列
Sum_buffer = {}
count_Alternative = float()

# '''替换nan'''
def replace_nan(str):
    if str == 'nan':
        return None
    return str
def replace_nan_float(str):
    if str == 'nan':
        return 0.0
    else:
        return float(str)

print('开始统计计划需求总量')
for _index,_row in excel_Plan_PurchaseBigTable.iterrows():
    branch=replace_nan(str(_row['子件编码']))
    branch_charge=replace_nan(str(_row['费用']))
    branch_substitution1=replace_nan(str(_row['替代料1']))
    branch_substitution2=replace_nan(str(_row['替代料2']))

    Branch1_List.append(branch)

    Branch = (branch,branch_charge,branch_substitution1,branch_substitution2)
    if (~(Branch_List.__contains__(Branch))):
        Branch_List.append(Branch)

    count_needsinNovember=replace_nan_float(str(_row[month_need]))
    if(Sum_plan.__contains__(branch)):
        Sum_plan[branch] += count_needsinNovember
    else:
        Sum_plan[branch] = count_needsinNovember


# 采购在途订单统计
print('开始统计采购在途数量')
for _index,_row in excel_PurchaseListTable.iterrows():
    branch=replace_nan(str(_row['存货编码']))
    count = replace_nan_float(str(_row['未入库量']))
    # branch_guanbiren=replace_nan(str(_row['关闭人']))
    branch_hangguanbiren=replace_nan(str(_row['行关闭人']))
    if branch_hangguanbiren==None:
        if(Sum_purchase.__contains__(branch)):
            Sum_purchase[branch] += count
        else:
            Sum_purchase[branch] = count

# 生产在制物料统计
print('开始统计生产在制数量')
for _index,_row in excel_MakingTable.iterrows():
    branch=replace_nan(str(_row['子件物料编码']))
    count_shouldget = replace_nan_float(str(_row['应领数量']))
    count_alreadyget = replace_nan_float(str(_row['已领数量']))
    count_noget = count_shouldget - count_alreadyget
    if(Sum_making.__contains__(branch)):
        Sum_making[branch] += count_noget
    else:
        Sum_making[branch] = count_noget

# 库存货位统计
print('开始统计库存数量')
for _index,_row in excel_StocksTable.iterrows():
    branch=replace_nan(str(_row['存货编码']))
    branch_store=replace_nan(str(_row['仓库名称']))
    count_nowstore = replace_nan_float(str(_row['现存数量']))
    if branch_store != '异常处理仓':
        if(Sum_stocks.__contains__(branch)):
            Sum_stocks[branch] += count_nowstore
        else:
            Sum_stocks[branch] = count_nowstore


# 所有在途及库存统计（含替代料）
print('在途物料汇总')
i = 0
Branch_head=[row[0] for row in Branch_List]
for Branch in Branch_List:
    count_Alternative = 0.0
    for branch in Branch:
        if ~(branch == None):
            if (Sum_purchase.__contains__(branch)):
                count_Alternative += Sum_purchase[branch]
            if(Sum_making.__contains__(branch)):
                count_Alternative += Sum_making[branch]
            if(Sum_stocks.__contains__(branch)):
                count_Alternative += Sum_stocks[branch]
    if(Sum_over.__contains__(Branch_head[i])):
        Sum_over[Branch_head[i]] += count_Alternative
    else:
        Sum_over[Branch_head[i]] = count_Alternative
    i += 1


# 缺料统计
print('开始统计缺料数量')
for branch in Sum_plan:
    if (Sum_over.__contains__(branch)):
        count_shortage = Sum_plan[branch] - Sum_over[branch]
    else:
        count_shortage = Sum_plan[branch]

    if(Sum_shortage.__contains__(branch)):
        Sum_shortage[branch] += count_shortage
    else:
        Sum_shortage[branch] = count_shortage


# 缺料数量 写入计划请购大表
print('正在写入中...')
# row_len = excel_Plan_PurchaseBigTable.shape[0]       # 行数
# column_len = excel_Plan_PurchaseBigTable.shape[1]    # 列数
row_len = sheet_NewPlan.max_row                    # 行数
column_len = sheet_NewPlan.max_column              # 列数

j = 0
sheet_NewPlan.cell(4,column_len+1,'缺件总数')
alignment_center = Alignment(horizontal='center', vertical='center')  # 居中
sheet_NewPlan.cell(4,column_len+1).alignment = alignment_center
sheet_NewPlan.cell(4,column_len+1).fill=PatternFill(patternType= 'solid',start_color=Color(index=5))

for branch in Branch1_List:
    sheet_NewPlan.cell(5+j,column_len+1).alignment = alignment_center
    sheet_NewPlan.cell(5+j,column_len+1).fill=PatternFill(patternType= 'solid',start_color=Color(index=5))
    if branch in Sum_shortage:
        sheet_NewPlan.cell(5+j,column_len+1,Sum_shortage[branch])
    else:
        sheet_NewPlan.cell(5+j,column_len+1,Sum_plan[branch])
    j += 1

newxl_path = Search_Path + '\生成大表.xlsx'
wb.save(newxl_path)
print('写入完成')
    # try:
    #     hostIDTable.objects.create(branch=branch,series=series,POdate=POdate,
    #     costNum1=costNum1,hostid=hostid,inventorycode=inventorycode,version=version,serialnumber=serialnumber,destination=destination,
    #     cliesnt=client,status=status,removestatus=removestatus,removedate=removedate,namedate=namedate,
    #     Amperage=Amperage,weight=weight,footmargin=footmargin,ionfan=ionfan,softlan=softlan,peculiarsetup=peculiarsetup,remark=remark,module=module,packagedate=packagedate,SEQdate=SEQdate,TEQdate=TEQdate,
    #     FEQdate=FEQdate,debugdate=debugdate,dedugenddate=dedugenddate,dutypeople=dutypeople,costnum2=costnum2,borrowdate=borrowdate,borrowname=borrowname)

    # except Exception as e :
    #     print (e)
print('123')
print('123')


