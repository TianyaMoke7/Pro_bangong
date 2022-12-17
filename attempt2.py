# 写EXCEL
# 储存：D:\Learning\attempt_file\EXCEL

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

new_path = 'D:\Learning\\attempt_file\EXCEL'   # 多加一个\，否则a被识别为0x07？？？？？？？？？？

wb = Workbook()

################################# 表1 ########################
ws1 = wb.active
ws1.title = '表1'

# ws1 = wb.add_sheet('表1')

ws1['A1'].value = '美国2021年国债发行量'

ws1.merge_cells('A1:D1')                        #合并单元格
# ws1.unmerge_cells('A1:D1')                    #拆分单元格
ws1.append([123,456,789,0])
ws1.append(['q','w','e','r'])


################################# 表2 ########################
data = [
    {
        'name':'小白',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name':'小黄',
        'tall': 177,
        'age': 28,
        'weight': 90
    },
    {
        'name':'小绿',
        'tall': 160,
        'age': 30,
        'weight': 60
    },
    {
        'name':'小灰',
        'tall': 155,
        'age': 50,
        'weight': 50
    },
    {
        'name':'小黑',
        'tall': 170,
        'age': 46,
        'weight': 99
    }
]

wb.create_sheet('表2')
ws2 = wb['表2']
ws2['A1'].value = '美国2020年国债发行量'
ws2.merge_cells('A1:D1')                        #合并单元格
ws2_title = ['姓名','身高','年龄','体重']
ws2.append(ws2_title)

for person in data:
    ws2.append(list(person.values()))

for col in range(2,5):
    char = get_column_letter(col)
    ws2[char + '8'] = f'=AVERAGE({char + "2"}:{char + "6"})'


newxl_path = new_path + '\美联储国债收益.xlsx'
wb.save(newxl_path)



        