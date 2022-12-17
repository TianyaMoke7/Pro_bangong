# 2022寒假作业升级版
# 加入无权限容错

from ast import While
import openpyxl
from openpyxl.styles import Font,colors, Alignment
from openpyxl import Workbook
from openpyxl.styles import *
# from  openpyxl import  Workbook
import xlrd
import xlwt
from pathlib import Path
import os
import shutil

from xlwt import Workbook
import string

import tkinter as tk
from tkinter import W, filedialog

import attempt10

root = tk.Tk()
root.withdraw()

Filepath = filedialog.askopenfilename() #获取选择好的文件
new_path = filedialog.askdirectory()  #获取选择好的文件夹



####################################### 开始读表 #########################
# 给出excel文件绝对路径
loc = str(Filepath) 
# wb_select = xlrd.open_workbook(loc)   # 在选定路径下打开工作表
wb_select = openpyxl.load_workbook(filename = loc) 

# sheet1_select = wb_select.sheets()[0]     #读取第一个sheet
sheet1_select = wb_select.worksheets[0] 
# print(sheet1_select.row_values(0))        #打印某一行数据
print(sheet1_select.cell(1,1).value)        #打印某一单元数据
print( )                    #打印空格

#print (wb.sheetnames) #出现报错
# sheet = wb_select.sheet_by_index(0)


############################ 遍历旧表1，提取关键字#######################
List1 = []
for i in range(sheet1_select.max_row):
     for j in range(sheet1_select.max_column):
         print(sheet1_select.cell(i+1, j+1).value)
         keyword = sheet1_select.cell(i+1, j+1).value
         List1.append(sheet1_select.cell(i+1, j+1).value)   #放入数列中，但数列把空格也计算在内了？？？？？？？？？？？？



##############################通过关键字在图库中查找图纸，然后把图纸放在选定文件夹下#######################

newxl_Vlue = {string:bool}

#folder = Path(Folderpath.strip())       #在选定文件夹下查找   ##### strip()去除首位空格

old_path_Dingzhi = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\15-定制品库\\'
old_path_08tuku = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\08-图库\\'
old_path_CunDang = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\12-U8数据查询\\003-金海通正式账套\\01-存货档案-每日更新-研发助理\\'
old_path_6000 = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\\6000制线图纸PDF版\\'
old_path_8000 = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\8000制线图纸PDF版\\'
old_path_ColliePlus = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\collie plus\ColliePLUS定制线束PDF\\'
# old_path_ColliePlus = '\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\collie plus\ColliePLUS\\'

# folder = Path(old_path_Dingzhi)                      #在定制品库中查找
Exist_6000 = os.path.exists(old_path_6000)
Exist_8000 = os.path.exists(old_path_8000)
Exist_ColliePlus = os.path.exists(old_path_ColliePlus)

file_names = os.listdir(old_path_Dingzhi)            #把定制品库中所有文件名放入列表
file_names_08tuku = os.listdir(old_path_08tuku)
file_names_CunDang = os.listdir(old_path_CunDang)
if Exist_6000:
    file_names_6000 = os.listdir(old_path_6000)
if Exist_8000:
    file_names_8000 = os.listdir(old_path_8000)
if Exist_ColliePlus:
    file_names_ColliePlus = os.listdir(old_path_ColliePlus)

#keyword = input("jh请输入要查找的文件或文件夹的名称：").strip()  #strip() 去除首尾空格

#在图库中查找List1中的关键字
# for keyword0 in List1:                     #在定制品库中匹配到ECD文件，匹配上后result长度为1，内容为字符串地址
#     result = list(folder.rglob(keyword0+'.pdf'))   ###此语句在公司电脑上贼费时间###    #list()将元组或字符串转换为数列。

#     if len(result) != 0:
#         print(f"在【{folder}】中查找到名为【{keyword0}】的文件或文件夹")
#         for file_address in result:
#             print(file_address)
#             #shutil.copyfile(os.path.join(file_address),os.path.join(new_path,file_name))
#     else:
#         print(f'在【{folder}】中未查找到名为【{keyword0}】的文件或文件夹！')
#         newxl_Vlue[keyword0] = False

#     for file_name in file_names:
#         if keyword0+'.pdf' == file_name:
#             #路径拼接要用os.path.join，复制指定文件到另一个文件夹里
#             shutil.copyfile(os.path.join(old_path_Dingzhi,file_name),os.path.join(new_path,file_name))


############################################### 打开存货档案 ####################################
for dangan in file_names_CunDang:
    if 'JHT 存货档案20' == dangan[0:10]:
        print('正在打开'+' “'+dangan+'” '+'请稍等...')
        break
dangan1 = old_path_CunDang + '\\' + dangan
excel_CunDang = xlrd.open_workbook(dangan1)
excel_CunDang_sheet1 = excel_CunDang.sheets()[0]    #第一个表单


################################################ 开始读6000文件查阅表
if Exist_6000:
    excel_6000 = xlrd.open_workbook(r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\00- 璞玉浑金\01-工艺文件\00-工艺储物柜\01-细分查阅表\6000制线文件查阅表.xlsx')
    # excel_6000 = openpyxl.load_workbook(filename = r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\00- 璞玉浑金\01-工艺文件\00-工艺储物柜\01-细分查阅表\6000制线文件查阅表.xlsx')
    # excel_6000_sheet1 = excel_6000.worksheets[0]       #第一个表单
    excel_6000_sheet1 = excel_6000.sheets()[0]       #第一个表单
    # excel_6000_sheet1_row_data = excel_6000_sheet1.row_values(1)  #第2行

if Exist_8000:
    excel_8000 = xlrd.open_workbook(r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\00- 璞玉浑金\01-工艺文件\00-工艺储物柜\01-细分查阅表\8000制线文件查阅表.xlsx')
    excel_8000_sheet1 = excel_8000.sheets()[0]

if Exist_ColliePlus:
    excel_ColliePlus = xlrd.open_workbook(r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\00- 璞玉浑金\01-工艺文件\00-工艺储物柜\01-细分查阅表\Collie Plus文件查阅表.xlsx')
    excel_ColliePlus_sheet1 = excel_ColliePlus.sheets()[0]

############################################### 创建表格并保存在新路径下 ################################

# 创建workbook
wb_establish = openpyxl.Workbook() 
  
# 使用add_sheet函数创建新的sheet，####写入首行

sheet1_establish = wb_establish.active
sheet1_establish.title = '定制品线缆'
sheet2_establish = wb_establish.create_sheet('6000线缆')
sheet3_establish = wb_establish.create_sheet('8000线缆')
sheet4_establish = wb_establish.create_sheet('Collie Plus线缆')
sheet5_establish = wb_establish.create_sheet('其他图纸')


################### 写入数据，参数分别为行、列、数据 

# 对格式进行操作

sheet1_establish.cell(1, 1, '图纸号')
sheet1_establish.cell(1, 2, '图纸名称') 
sheet1_establish.cell(1, 3, '规格型号')
sheet1_establish.cell(1, 4, 'pdf链接')
for j in range(1,5):
    sheet1_establish.cell(1, j).font = Font(bold=True)       #加粗
    sheet1_establish.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')   #居中
    sheet1_establish.cell(1, j).fill=PatternFill(patternType= 'solid',start_color=Color(index=44)) #背景颜色为灰蓝
sheet1_establish.column_dimensions['B'].width = 35
sheet1_establish.column_dimensions['C'].width = 25
sheet1_establish.column_dimensions['D'].width = 14


sheet2_establish.cell(1, 1, '图纸号')
sheet2_establish.cell(1, 2, '图纸名称') 
sheet2_establish.cell(1, 3, '版本')
sheet2_establish.cell(1, 4, 'pdf链接')
if ~Exist_6000:
    sheet2_establish.cell(2, 1, '无权查阅或6000图纸地址变更')
    sheet2_establish.cell(2, 4, '无权查阅或6000图纸地址变更')
for j in range(1,5):
    sheet2_establish.cell(1, j).font = Font(bold=True)
    sheet2_establish.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
    sheet2_establish.cell(1, j).fill=PatternFill(patternType= 'solid',start_color=Color(index=44))
sheet2_establish.column_dimensions['B'].width = 28
sheet2_establish.column_dimensions['D'].width = 21

sheet3_establish.cell(1, 1, '图纸号')
sheet3_establish.cell(1, 2, '图纸名称') 
sheet3_establish.cell(1, 3, '版本')
sheet3_establish.cell(1, 4, 'pdf链接')
if ~Exist_8000:
    sheet3_establish.cell(2, 1, '无权查阅或8000图纸地址变更')
    sheet3_establish.cell(2, 4, '无权查阅或8000图纸地址变更')
for j in range(1,5):
    sheet3_establish.cell(1, j).font = Font(bold=True)
    sheet3_establish.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
    sheet3_establish.cell(1, j).fill=PatternFill(patternType= 'solid',start_color=Color(index=44))
sheet3_establish.column_dimensions['B'].width = 28
sheet3_establish.column_dimensions['D'].width = 21

sheet4_establish.cell(1, 1, '图纸号')
sheet4_establish.cell(1, 2, '图纸名称') 
sheet4_establish.cell(1, 3, '版本')
sheet4_establish.cell(1, 4, 'pdf链接')
if ~Exist_ColliePlus:
    sheet4_establish.cell(2, 1, '无权查阅或ColliePlus图纸地址变更')
    sheet4_establish.cell(2, 4, '无权查阅或ColliePlus图纸地址变更')
for j in range(1,5):
    sheet4_establish.cell(1, j).font = Font(bold=True)
    sheet4_establish.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
    sheet4_establish.cell(1, j).fill=PatternFill(patternType= 'solid',start_color=Color(index=44))
sheet4_establish.column_dimensions['B'].width = 28
sheet4_establish.column_dimensions['D'].width = 21

sheet5_establish.cell(1, 1, '图纸号')
sheet5_establish.cell(1, 2, '图纸名称') 
sheet5_establish.cell(1, 3, '规格型号')
sheet5_establish.cell(1, 4, 'pdf链接')
sheet5_establish.cell(1, 5, '存放地址')
for j in range(1,6):
    sheet5_establish.cell(1, j).font = Font(bold=True)
    sheet5_establish.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')
    sheet5_establish.cell(1, j).fill=PatternFill(patternType= 'solid',start_color=Color(index=44))
sheet5_establish.column_dimensions['A'].width = 17
sheet5_establish.column_dimensions['B'].width = 17
sheet5_establish.column_dimensions['C'].width = 25
sheet5_establish.column_dimensions['D'].width = 25
sheet5_establish.column_dimensions['E'].width = 14


######################### 写入第一列，图纸号
i_ECD = 1
i_6000 = 1
i_8000 = 1
i_Collieplus = 1
i_others_Dingzhi = 1
i_08tuku = 1

for keyword0 in List1:
    keyword0_list = [keyword0]
    keyword1 = str(keyword0)
 
    ################################         ECD图纸查找结果       #####################
    if keyword1[0:3] == 'ECD':      
        #sheet1_establish.write(i, 0, keyword0) 
        sheet1_establish.append(keyword0_list)  #写入首列,每行追加
        i_ECD += 1

        # result_ECD = list(folder.rglob(keyword0+'.pdf'))     #在定制品库中去匹配ECD文件，匹配上后result长度为1，内容为字符串地址

        # for filename_ECD in file_names:
        #     if keyword0 == filename_ECD:
        #         shutil.copyfile(os.path.join(old_path_Dingzhi,keyword0+'.pdf'),os.path.join(new_path,keyword0+'.pdf'))
        #         sheet1_establish.cell(i_ECD, 3, '=HYPERLINK("\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\15-定制品库\{}","{}")\n'.strip().format(keyword0+'.pdf',keyword0+'.pdf'))
        #         print(f"在【{folder}】中查找到名为【{keyword0}】的文件或文件夹")
            

        if keyword0+'.pdf' in file_names:
            #路径拼接要用os.path.join，复制指定文件到另一个文件夹里
            shutil.copyfile(os.path.join(old_path_Dingzhi,keyword0+'.pdf'),os.path.join(new_path,keyword0+'.pdf'))

            sheet1_establish.cell(i_ECD, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\15-定制品库\{}","{}")\n'.strip().format(keyword0+'.pdf',keyword0+'.pdf'))
            
            print(f"在定制品库中查找到名为【{keyword0}】的文件或文件夹")

            if keyword0 in excel_CunDang_sheet1.col_values(2):     #遍历存货档案的第3列 存货编码
                row_num_CunDang = excel_CunDang_sheet1.col_values(2).index(keyword0)        #获取该单元格行号
                CunDang_sheet1_row_data = excel_CunDang_sheet1.row_values(row_num_CunDang)  #把该物料的一整行放入一个数组
                sheet1_establish.cell(i_ECD, 2, CunDang_sheet1_row_data[4])
                sheet1_establish.cell(i_ECD, 3, CunDang_sheet1_row_data[5])

        # if keyword0+'.pdf' not in file_names:
        else:
            sheet1_establish.cell(i_ECD, 4, '无图')
            sheet1_establish.cell(i_ECD, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
            print(f'在定制品库中未查找到名为【{keyword0}】的文件或文件夹！')



    ################################         6000图纸查找结果
    elif keyword1[0:3] == 'CC1' or keyword1[0:3] == 'CF1' or keyword1[0:3] == 'CG1' or keyword1[0:3] == 'CS1' or keyword1 == 'CJP0139':
        if Exist_6000:
            sheet2_establish.append(keyword0_list)  #写入首列
            i_6000 += 1

            if keyword0 in excel_6000_sheet1.col_values(1):     #遍历6000文件查阅表的第2列

                row_num_6000 = excel_6000_sheet1.col_values(1).index(keyword0)      #获取该单元格行号
                e6000_sheet1_row_data = excel_6000_sheet1.row_values(row_num_6000)  #把有图的一整行放入一个数组
                sheet2_establish.cell(i_6000, 1, e6000_sheet1_row_data[1]) 
                sheet2_establish.cell(i_6000, 2, e6000_sheet1_row_data[2])
                sheet2_establish.cell(i_6000, 3, e6000_sheet1_row_data[3])
                
                print(f"在6000制线文件查阅表中查找到【{keyword0}】")
                tuzhi_6000 = e6000_sheet1_row_data[0] +'.pdf'
                if tuzhi_6000 in file_names_6000:
                    sheet2_establish.cell(i_6000, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\\6000制线图纸PDF版\{}","{}")\n'.strip( ).format(e6000_sheet1_row_data[0] +'.pdf',e6000_sheet1_row_data[0] +'.pdf'))
                    shutil.copyfile(os.path.join(old_path_6000,tuzhi_6000),os.path.join(new_path,tuzhi_6000))
                else:
                    sheet2_establish.cell(i_6000, 4, '无图')
                    sheet2_establish.cell(i_6000, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
                
            else:
                sheet2_establish.cell(i_6000, 4, '无图')
                sheet2_establish.cell(i_6000, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
                print(f'在6000制线文件查阅表中未找到【{keyword0}】！')


    ################################         8000图纸查找结果
    elif keyword1[0:3] == 'CC2' or keyword1[0:3] == 'CF2' or keyword1[0:3] == 'CG2' or keyword1[0:3] == 'CS2' or keyword1 == 'CJP0145' or keyword1 == 'CJP0054' or keyword1 == 'BGBE002':
        if Exist_8000:
            sheet3_establish.append(keyword0_list)  #写入首列
            i_8000 += 1

            if keyword0 in excel_8000_sheet1.col_values(1):     #遍历8000文件查阅表的第2列
                
                row_num_8000 = excel_8000_sheet1.col_values(1).index(keyword0) #获取单元格行号
                e8000_sheet1_row_data = excel_8000_sheet1.row_values(row_num_8000)  #把有图的一整行放入一个数组
                sheet3_establish.cell(i_8000, 1, e8000_sheet1_row_data[1]) 
                sheet3_establish.cell(i_8000, 2, e8000_sheet1_row_data[2])
                sheet3_establish.cell(i_8000, 3, e8000_sheet1_row_data[3])
                
                print(f"在8000制线文件查阅表中查找到【{keyword0}】")
                tuzhi_8000 = e8000_sheet1_row_data[0] +'.pdf'

                if tuzhi_8000 in file_names_8000:
                    sheet3_establish.cell(i_8000, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\\8000制线图纸PDF版\{}","{}")\n'.strip( ).format(e8000_sheet1_row_data[0] +'.pdf',e8000_sheet1_row_data[0] +'.pdf'))
                    shutil.copyfile(os.path.join(old_path_8000,tuzhi_8000),os.path.join(new_path,tuzhi_8000))
                else:
                    sheet3_establish.cell(i_8000, 4, '无图')
                    sheet3_establish.cell(i_8000, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
                
            else:
                sheet3_establish.cell(i_8000, 4, '无图')
                sheet3_establish.cell(i_8000, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
                print(f'在8000制线文件查阅表中未找到【{keyword0}】！')



    ################################         Collie Plus图纸查找结果
    elif keyword1[0:3] == 'CC9' or keyword1[0:3] == 'CF9' or keyword1[0:3] == 'CG9' or keyword1[0:3] == 'CS9':
        if Exist_ColliePlus:
            sheet4_establish.append(keyword0_list)  #写入首列
            i_Collieplus += 1

            if keyword0 in excel_ColliePlus_sheet1.col_values(1):     #遍历ColliePlus文件查阅表的第2列
                
                row_num_ColliePlus = excel_ColliePlus_sheet1.col_values(1).index(keyword0) #获取单元格行号
                eColliePlus_sheet1_row_data = excel_ColliePlus_sheet1.row_values(row_num_ColliePlus)  #把有图的一整行放入一个数组
                sheet4_establish.cell(i_Collieplus, 1, eColliePlus_sheet1_row_data[1]) 
                sheet4_establish.cell(i_Collieplus, 2, eColliePlus_sheet1_row_data[2])
                sheet4_establish.cell(i_Collieplus, 3, eColliePlus_sheet1_row_data[3])
                
                print(f"在Collie Plus制线文件查阅表中查找到名为【{keyword0}】")
                tuzhi_ColliePlus = eColliePlus_sheet1_row_data[0] +'.pdf'

                if tuzhi_ColliePlus in file_names_ColliePlus:
                    sheet4_establish.cell(i_Collieplus, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\工艺部\\00- 璞玉浑金\\01-工艺文件\\00-工艺储物柜\\02-PDF\collie plus\ColliePlus\{}","{}")\n'.strip().format(eColliePlus_sheet1_row_data[0] +'.pdf',eColliePlus_sheet1_row_data[0] +'.pdf'))
                    shutil.copyfile(os.path.join(old_path_ColliePlus,tuzhi_ColliePlus),os.path.join(new_path,tuzhi_ColliePlus))
                else:
                    sheet4_establish.cell(i_Collieplus, 4, '无图')
                    sheet4_establish.cell(i_Collieplus, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))

            else:
                sheet4_establish.cell(i_Collieplus, 4, '无图')
                sheet4_establish.cell(i_Collieplus, 4).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))
                print(f'在Collie Plus制线文件查阅表中未找到【{keyword0}】！')


    ################################################   定制品库和08图库 查找结果  ###########################################
    else:
        sheet5_establish.append(keyword0_list)  #写入首列,每行追加
        i_others_Dingzhi += 1

        sheet5_establish.cell(i_others_Dingzhi, 5, '无图')
        sheet5_establish.cell(i_others_Dingzhi, 5).fill=PatternFill(patternType= 'solid',start_color=Color(index=10))

        for chart_Dingzhi in file_names:
            if keyword1 in chart_Dingzhi and len(keyword1) >= 7 and keyword1[0:3] != 'ECD':

                print(f"在定制品库中查找到名为【{keyword0}】的文件或文件夹")
                # lujing = str(old_path_Dingzhi1) +'\\' + chart_Dingzhi
                old_path_Dingzhi1 = r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\15-定制品库'+'\\'+chart_Dingzhi
                new_path1 = new_path.replace('/','\\')
                attempt10.fuzhi(old_path_Dingzhi1,new_path1)
                sheet5_establish.cell(i_others_Dingzhi, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\15-定制品库\{}","{}")\n'.strip().format(chart_Dingzhi,chart_Dingzhi))
                sheet5_establish.cell(i_others_Dingzhi, 5, '定制品库有图')
                sheet5_establish.cell(i_others_Dingzhi, 5).fill=PatternFill(patternType= 'solid',start_color=Color(index=11))
                
                if keyword0 in excel_CunDang_sheet1.col_values(2):     #遍历存货档案的第3列 存货编码
                    row_num_CunDang = excel_CunDang_sheet1.col_values(2).index(keyword0)        #获取该单元格行号
                    CunDang_sheet1_row_data = excel_CunDang_sheet1.row_values(row_num_CunDang)  #把该物料的一整行放入一个数组
                    sheet5_establish.cell(i_others_Dingzhi, 2, CunDang_sheet1_row_data[4])
                    sheet5_establish.cell(i_others_Dingzhi, 3, CunDang_sheet1_row_data[5])

        for chart_08tuku in file_names_08tuku:
            if keyword1 in chart_08tuku and len(keyword1) >= 7 and keyword1[0:3] != 'ECD':

                old_path_08tuku1 = r'\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\08-图库'+'\\'+chart_08tuku
                new_path1 = new_path.replace('/','\\')
                attempt10.fuzhi(old_path_08tuku1,new_path1)
                sheet5_establish.cell(i_others_Dingzhi, 4, '=HYPERLINK("\\\\172.16.1.52\JHT-design\JHT-CKP-CIP-文件管理中心\文件中心\\08-图库\{}","{}")\n'.strip().format(keyword0+'.pdf',keyword0+'.pdf'))
                sheet5_establish.cell(i_others_Dingzhi, 5, '08-图库有图')
                sheet5_establish.cell(i_others_Dingzhi, 5).fill=PatternFill(patternType= 'solid',start_color=Color(index=11))
                print(f"在08-图库中查找到名为【{keyword0}】的文件或文件夹")

                if keyword0 in excel_CunDang_sheet1.col_values(2):     #遍历存货档案的第3列 存货编码
                    row_num_CunDang = excel_CunDang_sheet1.col_values(2).index(keyword0)        #获取该单元格行号
                    CunDang_sheet1_row_data = excel_CunDang_sheet1.row_values(row_num_CunDang)  #把该物料的一整行放入一个数组
                    sheet5_establish.cell(i_others_Dingzhi, 2, CunDang_sheet1_row_data[4])
                    sheet5_establish.cell(i_others_Dingzhi, 3, CunDang_sheet1_row_data[5])
        # print(f'在所有图库中均未查找到名为【{keyword0}】的文件或文件夹！')  


# while True:
#     if i == len(List)+1:
#         break


#ws.write(i,1,"=HYPERLINK(\"#sheet2!a{}\")\r".format(i))   #添加超链接

#sheet1.write(1, 1, 'xlwt.Formula(u"HYPERLINK(\"D:\Learning\Projects\目标文件夹\Python学习路线图.pdf\",\"Python学习路线图.pdf\")\r"')

# 保存到excel表格
newxl_path = new_path + '\生成表格2.xlsx'
wb_establish.save(newxl_path)

